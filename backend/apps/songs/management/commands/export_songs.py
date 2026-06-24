import os
from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class Command(BaseCommand):
    help = 'Esporta i canti di un gruppo in un file Excel importabile.'

    def add_arguments(self, parser):
        parser.add_argument('--group', required=True, help='Nome o ID del gruppo da esportare')
        parser.add_argument('--out', default='export_canti.xlsx', help='Percorso file di output')

    def handle(self, *args, **options):
        from apps.songs.models import Song
        from apps.groups.models import MusicGroup

        group_arg = options['group']
        try:
            group = MusicGroup.objects.get(pk=group_arg)
        except (MusicGroup.DoesNotExist, Exception):
            try:
                group = MusicGroup.objects.get(name__iexact=group_arg)
            except MusicGroup.DoesNotExist:
                raise CommandError(f'Gruppo "{group_arg}" non trovato.')

        songs = Song.objects.filter(group=group).order_by('song_number', 'title')
        count = songs.count()
        if count == 0:
            self.stdout.write(self.style.WARNING('Nessun canto trovato nel gruppo.'))
            return

        wb = Workbook()
        ws = wb.active
        ws.title = 'Canti'

        HDR_FILL = PatternFill('solid', fgColor='1F4E79')
        R1_FILL  = PatternFill('solid', fgColor='EBF3FB')
        R2_FILL  = PatternFill('solid', fgColor='F5FAFF')
        thin = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        )

        headers = ['title', 'artist', 'key', 'mode', 'bpm', 'time_signature', 'notes', 'song_number', 'topics']
        widths  = [35, 25, 10, 12, 10, 16, 45, 14, 45]

        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(1, ci, h)
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = 'A2'

        for ri, song in enumerate(songs, 2):
            fill = R1_FILL if ri % 2 == 0 else R2_FILL
            topics_str = ', '.join(song.topics) if song.topics else ''
            row = [
                song.title,
                song.artist or '',
                song.key or '',
                song.mode or '',
                song.bpm or '',
                song.time_signature or '',
                song.notes or '',
                song.song_number or '',
                topics_str,
            ]
            for ci, val in enumerate(row, 1):
                c = ws.cell(ri, ci, val)
                c.font = Font(size=10)
                c.fill = fill
                c.border = thin
                c.alignment = Alignment(vertical='center', wrap_text=False)
            ws.row_dimensions[ri].height = 18

        ws.sheet_view.showGridLines = False

        out_path = options['out']
        wb.save(out_path)
        self.stdout.write(self.style.SUCCESS(
            f'Esportati {count} canti del gruppo "{group.name}" -> {os.path.abspath(out_path)}'
        ))
