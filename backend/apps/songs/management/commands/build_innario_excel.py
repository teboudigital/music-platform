"""
Genera backend/data/excel/base/innario_verso.xlsx
File Excel base per l'import dell'innario completo.
Uso: python manage.py build_innario_excel
"""
import os
from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


KEYS = ['C', 'C#', 'Db', 'D', 'D#', 'Eb', 'E', 'F', 'F#', 'Gb', 'G', 'G#', 'Ab', 'A', 'A#', 'Bb', 'B']

GENRES = [
    'Gospel', 'Spiritual', 'Inno', 'Canto liturgico', 'Contemporaneo',
    'Jazz', 'Blues', 'Soul', 'R&B', 'Pop', 'Rock', 'Folk', 'Classica',
    'Reggae', 'Afrobeat', 'Makossa', 'Bikutsi', 'Ndombolo', 'Soukous',
    'Amapiano', 'Highlife', 'Kizomba', 'Zouk', 'Salsa', 'Bachata',
    'Hip-hop', 'Rap', 'Trap', 'House', 'EDM', 'Altro',
]

# Colonne: (intestazione, larghezza, obbligatorio, descrizione)
COLUMNS = [
    ('Titolo',              38, True,  'Titolo del canto'),
    ('Autore',              26, False, 'Autore / compositore'),
    ('Interprete',          26, False, 'Interprete (se diverso dall\'autore)'),
    ('Genere',              20, False, 'Genere musicale (scegli dal menu)'),
    ('Genere personalizzato', 22, False, 'Genere non in lista'),
    ('Tonalità',            12, False, 'Tonalità (scegli dal menu)'),
    ('Modalità',            14, False, 'major = Maggiore  |  minor = Minore'),
    ('BPM',                 10, False, 'Battiti per minuto (numero intero)'),
    ('Ritmo',               12, False, 'Es: 4/4  3/4  6/8  12/8'),
    ('Testo',               55, False, 'Testo completo con accordi inline es: [Am]Venite'),
    ('Accordi',             30, False, 'Lista accordi separati da virgola es: Am, F, C/E, G'),
    ('Note',                40, False, 'Note aggiuntive libere'),
    ('N° Innario',          14, False, 'Numero nel libretto ufficiale'),
    ('Argomenti',           35, False, 'Categorie separate da virgola es: adorazione, lode'),
]


class Command(BaseCommand):
    help = 'Genera il file Excel base innario_verso.xlsx in data/excel/base/'

    def handle(self, *args, **options):
        out_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            '..', '..', '..', '..', 'data', 'excel', 'base', 'innario_verso.xlsx'
        )
        out_path = os.path.normpath(out_path)

        wb = Workbook()

        # ── Stili condivisi ───────────────────────────────────────────────────
        thin = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF'),
        )

        HDR_FILL_REQ = PatternFill('solid', fgColor='1F4E79')
        HDR_FILL_OPT = PatternFill('solid', fgColor='2E75B6')
        ROW_FILL_REQ = PatternFill('solid', fgColor='FFF2F2')
        ROW_FILL_OPT_A = PatternFill('solid', fgColor='EBF3FB')
        ROW_FILL_OPT_B = PatternFill('solid', fgColor='F5FAFF')

        # ── Foglio 1: Istruzioni ──────────────────────────────────────────────
        ws_help = wb.active
        ws_help.title = 'Istruzioni'
        ws_help.sheet_view.showGridLines = False
        ws_help.column_dimensions['A'].width = 26
        ws_help.column_dimensions['B'].width = 52
        ws_help.column_dimensions['C'].width = 32

        ws_help.merge_cells('A1:C1')
        t = ws_help.cell(1, 1, '🎵  Innario — Guida all\'importazione')
        t.font = Font(bold=True, size=15, color='1F4E79')
        t.alignment = Alignment(horizontal='left', vertical='center')
        ws_help.row_dimensions[1].height = 32

        ws_help.merge_cells('A2:C2')
        sub = ws_help.cell(2, 1, 'Compila il foglio "Innario" e importalo dall\'app. I campi con ★ sono obbligatori.')
        sub.font = Font(size=10, italic=True, color='595959')
        ws_help.row_dimensions[2].height = 18

        ws_help.row_dimensions[3].height = 8

        head_fill = PatternFill('solid', fgColor='1F4E79')
        for ci, label in enumerate(['Colonna', 'Descrizione', 'Valori / Formato'], 1):
            c = ws_help.cell(4, ci, label)
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.fill = head_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin
        ws_help.row_dimensions[4].height = 22

        value_hints = [
            'Testo libero',
            'Testo libero',
            'Testo libero',
            'Scegli dal menu a tendina',
            'Testo libero (se genere non in lista)',
            '  |  '.join(KEYS),
            'major  oppure  minor',
            'Numero intero es: 120',
            'es: 4/4  3/4  6/8  12/8',
            '[Am]Venite ado[F]riamo il [C]Signore',
            'Am, F, C/E, G  (virgola tra accordi)',
            'Testo libero',
            'Numero intero es: 42',
            'adorazione, lode, Maria  (virgola)',
        ]

        req_bg = PatternFill('solid', fgColor='FFF2F2')
        opt_bg = PatternFill('solid', fgColor='EBF3FB')

        for r, ((col_name, _, required, desc), hint) in enumerate(zip(COLUMNS, value_hints), start=5):
            bg = req_bg if required else opt_bg
            label = f'{col_name}  ★' if required else col_name
            for ci, val in enumerate([label, desc, hint], 1):
                c = ws_help.cell(r, ci, val)
                c.font = Font(bold=(ci == 1 and required), size=10)
                c.fill = bg
                c.border = thin
                c.alignment = Alignment(vertical='center', wrap_text=True)
            ws_help.row_dimensions[r].height = 28

        note_row = 5 + len(COLUMNS) + 1
        ws_help.merge_cells(f'A{note_row}:C{note_row}')
        note = ws_help.cell(note_row, 1, '→ Vai al foglio "Innario" per inserire i dati')
        note.font = Font(bold=True, color='2E75B6', size=11)

        # ── Foglio 2: Innario ─────────────────────────────────────────────────
        ws = wb.create_sheet('Innario')
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A2'

        for ci, (name, width, required, _) in enumerate(COLUMNS, 1):
            col_letter = get_column_letter(ci)
            label = f'{name}  ★' if required else name
            c = ws.cell(1, ci, label)
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.fill = HDR_FILL_REQ if required else HDR_FILL_OPT
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin
            ws.column_dimensions[col_letter].width = width

        ws.row_dimensions[1].height = 28

        for ri in range(2, 601):
            fill_req = ROW_FILL_REQ
            fill_opt = ROW_FILL_OPT_A if ri % 2 == 0 else ROW_FILL_OPT_B
            for ci, (_, _, required, _) in enumerate(COLUMNS, 1):
                c = ws.cell(ri, ci)
                c.fill = fill_req if required else fill_opt
                c.border = thin
                c.font = Font(size=10)
                # Testo (col 10) e Accordi (col 11) e Note (col 12) con wrap
                c.alignment = Alignment(vertical='top', wrap_text=(ci in (10, 11, 12, 14)))
            ws.row_dimensions[ri].height = 18

        # Dropdown Tonalità (col F = 6)
        dv_key = DataValidation(
            type='list',
            formula1='"' + ','.join(KEYS) + '"',
            allow_blank=True,
            showDropDown=False,
            error='Scegli dal menu.',
            errorTitle='Tonalità non valida',
            prompt='Scegli la tonalità',
            promptTitle='Tonalità',
        )
        dv_key.sqref = f'{get_column_letter(6)}2:{get_column_letter(6)}600'
        ws.add_data_validation(dv_key)

        # Dropdown Modalità (col G = 7)
        dv_mode = DataValidation(
            type='list',
            formula1='"major,minor"',
            allow_blank=True,
            showDropDown=False,
            error='Scrivi "major" o "minor".',
            errorTitle='Modalità non valida',
            prompt='major = Maggiore  |  minor = Minore',
            promptTitle='Modalità',
        )
        dv_mode.sqref = f'{get_column_letter(7)}2:{get_column_letter(7)}600'
        ws.add_data_validation(dv_mode)

        # Dropdown Genere (col D = 4)
        genres_str = ','.join(GENRES)
        dv_genre = DataValidation(
            type='list',
            formula1='"' + genres_str + '"',
            allow_blank=True,
            showDropDown=False,
            error='Scegli dal menu o usa "Genere personalizzato".',
            errorTitle='Genere non valido',
            prompt='Scegli il genere musicale',
            promptTitle='Genere',
        )
        dv_genre.sqref = f'{get_column_letter(4)}2:{get_column_letter(4)}600'
        ws.add_data_validation(dv_genre)

        # Riga esempio
        example = [
            'Ave Maria', 'Franz Schubert', '', 'Classica', '',
            'F', 'major', 120, '4/4',
            '[F]Ave Ma[C]ria gra[Dm]tia ple[C]na',
            'F, C, Dm, Bb, C/E',
            'Versione in Fa maggiore',
            1, 'adorazione, Maria',
        ]
        for ci, val in enumerate(example, 1):
            c = ws.cell(2, ci, val)
            c.font = Font(italic=True, color='595959', size=9)

        wb.active = ws
        wb.save(out_path)
        self.stdout.write(self.style.SUCCESS(f'File generato: {out_path}'))
